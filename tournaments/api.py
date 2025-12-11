from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
import pyotp
from rest_framework import status
from django.utils import timezone
from datetime import timedelta
import random
from rest_framework.viewsets import GenericViewSet
from rest_framework import mixins, viewsets, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import serializers
from django.db.models import Q, Count, Avg, Max, Min
from django.contrib.auth import authenticate, login, logout

from .models import Team, Player, Tournament, Match, TournamentCategory
from .serializers import (
    TeamsCRSerializer, TeamsUDSerializer,
    PlayersCRSerializer, PlayersUDSerializer,
    TournamentsCRSerializer, TournamentsUDSerializer,
    MatchesCRSerializer, MatchesUDSerializer,
    TournamentCategoriesCRSerializer, TournamentCategoriesUDSerializer
)

class UserViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]

    @action(detail=False, url_path="info", methods=["GET"])
    def get_info(self, request, *args, **kwargs):
       
        user_info = {
            "username": request.user.username,
            "is_authenticated": request.user.is_authenticated,
            "is_staff": request.user.is_staff,
      
        }
        if self.request.user.is_authenticated:
            user_info.update({
                'second': self.request.session.get('second') or False, 
            })
        
        
        return Response(user_info)

    @action(detail=False, url_path="login", methods=["POST"])
    def login_user(self, request, *args, **kwargs):
        username = self.request.data['username']
        password = self.request.data['password']

        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            return Response({
                'success': True,
            })

        return Response({
            'success': False,
        })
    
    @action(detail=False, url_path="second-login", methods=["POST"])
    def second_login(self, *args, **kwargs):
        player = Player.objects.get(user=self.request.user) 
        key = player.totp_key  
        t = pyotp.totp.TOTP(key)
        input_code = self.request.data.get('key', '')
        if input_code == t.now():
            self.request.session['second'] = True
            return Response({
                'success': True,
            })

    @action(detail=False, url_path="logout", methods=["POST"])
    def logout_user(self, request, *args, **kwargs):
        request.session.flush()
        logout(request)
        return Response({
            'success': True,
        }, status=status.HTTP_200_OK)

class TeamsViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin, 
    GenericViewSet
):
    queryset = Team.objects.all()
    serializer_class = TeamsCRSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update']:
            return TeamsUDSerializer
        return super().get_serializer_class()

    class TeamStatsSerializer(serializers.Serializer):
        total_teams = serializers.IntegerField()
        teams_with_players = serializers.IntegerField()
        teams_without_players = serializers.IntegerField()
        avg_players_per_team = serializers.FloatField()
        max_players_in_team = serializers.IntegerField()
        min_players_in_team = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_team_stats(self, request, *args, **kwargs):
        team_stats = Team.objects.annotate(
            player_count=Count('player')
        ).aggregate(
            total_teams=Count("*"),
            teams_with_players=Count("id", filter=Q(player_count__gt=0)),
            teams_without_players=Count("id", filter=Q(player_count=0)),
            avg_players_per_team=Avg("player_count"),
            max_players_in_team=Max("player_count"),
            min_players_in_team=Min("player_count"),
        )
        
        stats = {
            'total_teams': team_stats['total_teams'],
            'teams_with_players': team_stats['teams_with_players'],
            'teams_without_players': team_stats['teams_without_players'],
            'avg_players_per_team': round(team_stats['avg_players_per_team'] or 0, 1),
            'max_players_in_team': team_stats['max_players_in_team'] or 0,
            'min_players_in_team': team_stats['min_players_in_team'] or 0,
        }
        
        serializer = self.TeamStatsSerializer(instance=stats)
        return Response(serializer.data)

class PlayersViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin, 
    GenericViewSet
):
    queryset = Player.objects.all()
    serializer_class = PlayersCRSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update']:
            return PlayersUDSerializer
        return super().get_serializer_class()

    class PlayerStatsSerializer(serializers.Serializer):
        total_players = serializers.IntegerField()
        players_with_team = serializers.IntegerField()
        players_without_team = serializers.IntegerField()
        players_with_user = serializers.IntegerField()
        players_without_user = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_player_stats(self, request, *args, **kwargs):
        player_stats = Player.objects.aggregate(
            total_players=Count("*"),
            players_with_team=Count("id", filter=Q(team__isnull=False)),
            players_without_team=Count("id", filter=Q(team__isnull=True)),
            players_with_user=Count("id", filter=Q(user__isnull=False)),
            players_without_user=Count("id", filter=Q(user__isnull=True)),
        )
        
        stats = {
            'total_players': player_stats['total_players'],
            'players_with_team': player_stats['players_with_team'],
            'players_without_team': player_stats['players_without_team'],
            'players_with_user': player_stats['players_with_user'],
            'players_without_user': player_stats['players_without_user'],
        }
        
        serializer = self.PlayerStatsSerializer(instance=stats)
        return Response(serializer.data)

class TournamentCategoriesViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin, 
    GenericViewSet
):
    queryset = TournamentCategory.objects.all()
    serializer_class = TournamentCategoriesCRSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update']:
            return TournamentCategoriesUDSerializer
        return super().get_serializer_class()

    class CategoryStatsSerializer(serializers.Serializer):
        total_categories = serializers.IntegerField()
        categories_with_tournaments = serializers.IntegerField()
        categories_without_tournaments = serializers.IntegerField()
        most_popular_category = serializers.CharField()
        tournaments_in_popular_category = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_category_stats(self, request, *args, **kwargs):
        category_stats = TournamentCategory.objects.annotate(
            tournament_count=Count('tournament')
        ).order_by('-tournament_count')
        
        total_categories = category_stats.count()
        categories_with_tournaments = category_stats.filter(tournament_count__gt=0).count()
        categories_without_tournaments = total_categories - categories_with_tournaments
        
        most_popular = category_stats.first()
        
        stats = {
            'total_categories': total_categories,
            'categories_with_tournaments': categories_with_tournaments,
            'categories_without_tournaments': categories_without_tournaments,
            'most_popular_category': most_popular.name if most_popular else "Нет данных",
            'tournaments_in_popular_category': most_popular.tournament_count if most_popular else 0,
        }
        
        serializer = self.CategoryStatsSerializer(instance=stats)
        return Response(serializer.data)

class TournamentsViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin, 
    GenericViewSet
):
    queryset = Tournament.objects.all()
    serializer_class = TournamentsCRSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update']:
            return TournamentsUDSerializer
        return super().get_serializer_class()

    class TournamentStatsSerializer(serializers.Serializer):
        total_tournaments = serializers.IntegerField()
        active_tournaments = serializers.IntegerField()
        completed_tournaments = serializers.IntegerField()
        upcoming_tournaments = serializers.IntegerField()
        avg_tournament_duration = serializers.FloatField()
        tournaments_with_matches = serializers.IntegerField()
        tournaments_without_matches = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_tournament_stats(self, request, *args, **kwargs):
        from django.utils import timezone
        
        now = timezone.now().date()
        
        tournament_stats = Tournament.objects.annotate(
            match_count=Count('match'),
            duration=Max('end_date') - Min('start_date')
        ).aggregate(
            total_tournaments=Count("*"),
            active_tournaments=Count("id", filter=Q(start_date__lte=now, end_date__gte=now)),
            completed_tournaments=Count("id", filter=Q(end_date__lt=now)),
            upcoming_tournaments=Count("id", filter=Q(start_date__gt=now)),
            tournaments_with_matches=Count("id", filter=Q(match_count__gt=0)),
            tournaments_without_matches=Count("id", filter=Q(match_count=0)),
        )
        
        # Средняя продолжительность турниров в днях
        tournaments_with_dates = Tournament.objects.filter(
            start_date__isnull=False, 
            end_date__isnull=False
        )
        if tournaments_with_dates.exists():
            avg_duration = tournaments_with_dates.aggregate(
                avg_days=Avg(Max('end_date') - Min('start_date'))
            )['avg_days']
            avg_duration_days = avg_duration.days if avg_duration else 0
        else:
            avg_duration_days = 0
        
        stats = {
            'total_tournaments': tournament_stats['total_tournaments'],
            'active_tournaments': tournament_stats['active_tournaments'],
            'completed_tournaments': tournament_stats['completed_tournaments'],
            'upcoming_tournaments': tournament_stats['upcoming_tournaments'],
            'avg_tournament_duration': round(avg_duration_days, 1),
            'tournaments_with_matches': tournament_stats['tournaments_with_matches'],
            'tournaments_without_matches': tournament_stats['tournaments_without_matches'],
        }
        
        serializer = self.TournamentStatsSerializer(instance=stats)
        return Response(serializer.data)

class MatchesViewSet(
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin, 
    GenericViewSet
):
    queryset = Match.objects.all()
    serializer_class = MatchesCRSerializer
    
    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]
    
    def get_serializer_class(self):
        if self.action in ['create', 'update']:
            return MatchesUDSerializer
        return super().get_serializer_class()
    
    def get_queryset(self):
        """
        Переопределяем queryset для фильтрации матчей:
        - Администраторы видят все матчи
        - Игроки видят только матчи своей команды
        - Неавторизованные пользователи видют все матчи
        """
        queryset = super().get_queryset()
        
        # Если пользователь администратор или неавторизован - возвращаем все матчи
        if not self.request.user.is_authenticated or self.request.user.is_staff:
            return queryset
        
        # Для авторизованных игроков фильтруем матчи по их команде
        try:
            player = Player.objects.get(user=self.request.user)
            if player.team:
                # Фильтруем матчи, где команда игрока участвует как team1 или team2
                queryset = queryset.filter(
                    Q(team1=player.team) | Q(team2=player.team)
                )
            else:
                # Если у игрока нет команды, возвращаем пустой queryset
                queryset = queryset.none()
        except Player.DoesNotExist:
            # Если пользователь не является игроком, возвращаем пустой queryset
            queryset = queryset.none()
        
        return queryset

    class MatchStatsSerializer(serializers.Serializer):
        total_matches = serializers.IntegerField()
        tournament_matches = serializers.IntegerField()
        non_tournament_matches = serializers.IntegerField()
        matches_with_winner = serializers.IntegerField()
        draws = serializers.IntegerField()
        avg_team1_score = serializers.FloatField()
        avg_team2_score = serializers.FloatField()
        highest_scoring_match = serializers.IntegerField()
        
    @action(detail=False, methods=["GET"], url_path="stats")
    def get_match_stats(self, request, *args, **kwargs):
        match_stats = Match.objects.aggregate(
            total_matches=Count("*"),
            tournament_matches=Count("id", filter=Q(tournament__isnull=False)),
            non_tournament_matches=Count("id", filter=Q(tournament__isnull=True)),
            matches_with_winner=Count("id", filter=Q(winner__isnull=False)),
            draws=Count("id", filter=Q(winner__isnull=True)),
            avg_team1_score=Avg("team1_score"),
            avg_team2_score=Avg("team2_score"),
            highest_scoring_match=Max("team1_score") + Max("team2_score"),
        )
        
        stats = {
            'total_matches': match_stats['total_matches'],
            'tournament_matches': match_stats['tournament_matches'],
            'non_tournament_matches': match_stats['non_tournament_matches'],
            'matches_with_winner': match_stats['matches_with_winner'],
            'draws': match_stats['draws'],
            'avg_team1_score': round(match_stats['avg_team1_score'] or 0, 1),
            'avg_team2_score': round(match_stats['avg_team2_score'] or 0, 1),
            'highest_scoring_match': match_stats['highest_scoring_match'] or 0,
        }
        
        serializer = self.MatchStatsSerializer(instance=stats)
        return Response(serializer.data)

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_matches_to_excel(self, request, *args, **kwargs):
        # Получаем матчи с предварительной загрузкой связанных данных
        matches = Match.objects.select_related(
            'tournament', 'team1', 'team2', 'winner'
        ).all()
       
        wb = Workbook()
        ws = wb.active
        ws.title = "Матчи турнирной системы"
        
        # Заголовки таблицы
        headers = [
            'ID матча',
            'Турнир', 
            'Команда 1', 
            'Счёт команды 1', 
            'Команда 2', 
            'Счёт команды 2',
            'Победитель',
            'Дата матча',
            'Статус'
        ]
        
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Заполнение данных
        for row, match in enumerate(matches, 2):
            # Определяем победителя
            winner_name = "Ничья"
            if match.winner:
                winner_name = match.winner.name
            
            # Определяем статус матча
            status = "Завершён" if match.winner else "В процессе"
            
            ws.cell(row=row, column=1, value=match.id)
            ws.cell(row=row, column=2, value=match.tournament.name if match.tournament else "Без турнира")
            ws.cell(row=row, column=3, value=match.team1.name if match.team1 else "Не указана")
            ws.cell(row=row, column=4, value=match.team1_score or 0)
            ws.cell(row=row, column=5, value=match.team2.name if match.team2 else "Не указана")
            ws.cell(row=row, column=6, value=match.team2_score or 0)
            ws.cell(row=row, column=7, value=winner_name)
            ws.cell(row=row, column=8, value=match.match_date.strftime('%d.%m.%Y %H:%M') if match.match_date else "Не указана")
            ws.cell(row=row, column=9, value=status)
        
        
        for column in ws.columns: 
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="tournament_matches.xlsx"'
        
        return response